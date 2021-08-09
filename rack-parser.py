#!/usr/bin/env python3
# Rack parser to flat file

import re
import sys
import json
import time
import argparse
from openpyxl import load_workbook

JUMP_LEFT_SEQ = '\u001b[100D'

# Search string for rack id (format: AA1.DC1.NN)
# AA - Address
# DC - Data center
# NN - Rack number
rack_id_regex = r'([A-Z][A-Z]\d)\.([A-Z][A-Z]\d)\.(\w{1,3})$'
ignore_list = [ r'^fc$', r'^lc.\w\w', r'^fc.\w\w', r'^fc\d+', r'^pdu\d*', r'^smu$', r'^cmu$', r'^lisa$',
                'empty', 'utp', 'reserve', 'organizer', 'service unit', 'pp-mm', 'patch',
                'shelf', 'tray', 'пусто', 'волс', 'патч', 'органайзер', 'полка', 'крс']
address_book = {}

start = time.monotonic()

parser = argparse.ArgumentParser(description='Converts rack unit view to flat inventory file')
parser.add_argument('source', type=argparse.FileType('rb'),
                    help='Rack view XLSX file')
parser.add_argument('-x', '--row', default=1000, metavar='X', type=int,
                    help='Maximum row to scan (default 1000)')
parser.add_argument('-y', '--column', default=1000, metavar='Y', type=int,
                    help='Maximum column to scan (default 1000)')
parser.add_argument('-b', '--buffer', default=100, metavar='N', type=int,
                    help='Break after N empty cells (default 100)')
parser.add_argument('-a', '--addr', metavar='json', type=argparse.FileType('r'),
                    help='Address book JSON file')
parser.add_argument('-v', '--verbose', help='More information', action='store_true')
args = parser.parse_args()

def is_num(n):
    """Checks value is an integer"""
    try:
        if n and n is not True:
            int(n)
            return True
        else: return False
    except ValueError:
        return False

def is_merge(row, column):
    """Checks cell is merged"""
    cell = ws.cell(row, column)
    for merged_cell in ws.merged_cells.ranges:
        if (cell.coordinate in merged_cell):
            return True
    return False

def bottom_border(x, y, size):
    """Checks that the bottom border is found"""
    # First merged cell has a border of whole range, so it need to be ignored
    if is_merge(x, y) is True and size == 1:
        return False
    # Checking bottom border of the current cell and top border of the cell under current one
    elif ws.cell(row=x, column=y).border.bottom.style or ws.cell(row=x+1, column=y).border.top.style:
        return True
    else:
        return False

def search_rack(rack_x, rack_y, progress):
    """Searching devices from rack id coordinates"""
    for x in range(rack_x, rack_x+60):
        value = ws.cell(row=x, column=rack_y-1).value
        # Finding units count
        if is_num(value):
            rack_unit = value
            label = get_label(x, rack_y)
            if label:
                vendor = get_info(x, rack_y+1, label['size'])
                model = get_info(x, rack_y+2, label['size'])
                serial = get_info(x, rack_y+3, label['size'])
                dev = prepare_device(vendor, model, serial, label, rack_id, rack_unit, progress)
                if dev:
                    csv = data_center + ';' + address + ';' + dev['model'] + ';' + dev['serial'] + ';' \
                        + label['name'] + ';' + str(rack_num) + ';' + str(rack_unit) + ';' + str(label['size'])
                    progress['devices'] += 1
                    if args.verbose:
                        print(f"{progress['devices']}.{csv}")
                    else:
                        print(JUMP_LEFT_SEQ, end='')
                        print(f"Processing: {rack_id} (racks: {progress['racks']}, found: {progress['devices']}, "
                              f"ignored: {progress['ignored']})", end='')
                        sys.stdout.flush()
                    with open('result.csv', 'a') as result:
                        result.write(f'{csv}\n')
            # Stoping on last unit
            if int(value) == 1: break 

def get_label(x, y):
    """Getting label name and size"""
    # Checks that unit has top border in label or bottom border above, this means the device starts here
    if ws.cell(row=x, column=y).border.top.style or (ws.cell(row=x-1, column=y).border.bottom.style and not is_merge(x-1, y)):
        label = {'size': 1, 'name': ''}
        for x in range(x, x+40):
            value = ws.cell(row=x, column=y).value
            if value:
                label['name'] += str(value).strip().replace('\n', ' ').replace(';',' ')
            if not bottom_border(x, y, label['size']): 
                label['size'] += 1
            else: 
                break
        return label
    else:
        return False

def get_info(x, y, unit_size):
    """Getting label attributes"""
    for x in range(x, x+unit_size):
        value = ws.cell(row=x, column=y).value
        if value:
            return str(value).strip().replace('\n', ' ').replace(';',' ')
    return False

def prepare_device(vendor, model, serial, label, rack_id, rack_unit, progress):
    """Preparing the device attributes output"""
    if vendor or model or serial or label['name']:
        if not vendor: vendor = ''
        if not model: model = ''
        if not serial: serial = ''
        for stop_word in ignore_list:
            for info in (vendor, model, label['name']):
                if re.search(stop_word, info, re.IGNORECASE) and not serial:
                    progress['ignored'] += 1
                    csv = rack_id + ';' + (vendor + ' ' + model).strip() + ';' \
                        + label['name'] + ';' + str(rack_unit) + ';' + str(label['size'])
                    if args.verbose:
                        print(f"IGNORED: {progress['ignored']}.{csv}")
                    with open('ignore.csv', 'a') as ignore:
                        ignore.write(f'{csv}\n')
                    return False
        rack_device = {}
        if vendor.islower(): vendor = vendor.capitalize()
        rack_device['model'] = (vendor + ' ' + model).strip()
        rack_device['serial'] = serial
        return rack_device
    else:
        return False

def set_address(addr_id):
    """Setting address from dictionary"""
    for id, address in address_book.items():
        if id == addr_id:
            return address
    return ''


if args.addr:
    with open(args.addr.name) as json_file:
        address_book = json.load(json_file)

# Clearing output files
open('result.csv', 'w').close()
open('ignore.csv', 'w').close()
# Writing headers
with open('result.csv', 'a') as result:
    result.write('Площадка;Адрес;Модель;S/N;Label;Стойка;Место в стойке;Кол-во юнитов\n')
with open('ignore.csv', 'a') as ignore:
    ignore.write('Идентификатор стойки;Модель;Label;Место в стойке;Кол-во юнитов\n')

wb = load_workbook(args.source)

progress = {'devices': 0, 'racks': 0, 'ignored': 0}

for ws in wb:
    # Reading all worksheets in book
    #print(ws.title)
    break_count_row = 1
    for x in range(1, args.row):
        row_not_empty = False
        break_count_col = 1
        for y in range(1, args.column):
            value = ws.cell(row=x, column=y).value
            if value:
                row_not_empty = True
                # Search for rack id
                output = re.match(rack_id_regex, str(value))
                if output:
                    progress['racks'] += 1
                    data_center = 'ЦОД-' + output.group(1) + '_' + output.group(2)
                    address = set_address(output.group(1))
                    rack_num = output.group(3)
                    rack_id = output.group(0)
                    search_rack(x, y, progress)
            else:
                break_count_col += 1
                if break_count_col > args.buffer: break
        # Checks for an empty stop buffer
        break_count_row += 1
        if row_not_empty: break_count_row = 1
        if break_count_row > args.buffer: break

elapsed = round(time.monotonic() - start)
m, s = divmod(elapsed, 60)
print(f'\nElapsed time: {m}m {s}s')